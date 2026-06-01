from __future__ import annotations

from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from openpyxl import load_workbook

from review.models import InDepthArea, InDepthStatement


# Keywords found in column headers → StandardType
# Checked in order; first match wins.
_HEADER_KEYWORD_MAP = [
    ("urgent improvement", InDepthStatement.StandardType.URGENT_IMPROVEMENT),
    ("urgent", InDepthStatement.StandardType.URGENT_IMPROVEMENT),
    ("needs attention", InDepthStatement.StandardType.NEEDS_ATTENTION),
    ("below expected", InDepthStatement.StandardType.NEEDS_ATTENTION),
    ("exceptional", InDepthStatement.StandardType.EXCEPTIONAL),
    ("strong standard", InDepthStatement.StandardType.STRONG_STANDARD),
    ("strong", InDepthStatement.StandardType.STRONG_STANDARD),
    # "Statement Text" column is the expected standard
    ("statement text", InDepthStatement.StandardType.EXPECTED),
    ("expected standard", InDepthStatement.StandardType.EXPECTED),
    ("expected", InDepthStatement.StandardType.EXPECTED),
]

# Cell values that mean "no statement for this level"
_SKIP_VALUES = {"not applicable", "n/a", "na", "-", ""}


def _header_to_standard_type(header: str):
    """Return a StandardType if the column header represents a statement level, else None."""
    lower = header.lower()
    for keyword, stype in _HEADER_KEYWORD_MAP:
        if keyword in lower:
            return stype
    return None


def _is_skippable(value) -> bool:
    if value is None:
        return True
    text = str(value).strip().lower()
    return text in _SKIP_VALUES or text.startswith("not applicable")


class Command(BaseCommand):
	help = "Import / update in-depth review statements from the Ofsted Excel file (wide format)."

	def add_arguments(self, parser):
		parser.add_argument(
			"--path",
			dest="path",
			default="data/ofsted_expected_standard_review_statements.xlsx",
			help="Path to the xlsx file (default: data/ofsted_expected_standard_review_statements.xlsx)",
		)
		parser.add_argument(
			"--sheet",
			dest="sheet",
			default=None,
			help="Worksheet name (default: first sheet containing the header row)",
		)

	def handle(self, *args, **options):
		path = Path(options["path"])
		if not path.exists():
			raise CommandError(f"File not found: {path}")

		wb = load_workbook(path, read_only=True)

		# Pick the sheet
		sheet_name = options["sheet"]
		if sheet_name:
			if sheet_name not in wb.sheetnames:
				raise CommandError(f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}")
			ws = wb[sheet_name]
		else:
			ws = wb[wb.sheetnames[0]]

		max_row = ws.max_row or 10000
		max_col = ws.max_column or 20

		# ── Find the header row ──────────────────────────────────────────────
		# Look for a row where col A = "Evaluation Area" and col B = "Statement Number"
		header_row = None
		for r in range(1, min(20, max_row) + 1):
			a = ws.cell(r, 1).value
			b = ws.cell(r, 2).value
			if (
				isinstance(a, str) and "evaluation area" in a.lower()
				and isinstance(b, str) and "statement" in b.lower()
			):
				header_row = r
				break

		if header_row is None:
			raise CommandError(
				"Could not find header row containing 'Evaluation Area' and 'Statement Number'."
			)

		# ── Map columns to standard types ────────────────────────────────────
		# col 1 = area, col 2 = statement number; remaining cols checked by header text.
		type_columns: list[tuple[int, InDepthStatement.StandardType]] = []
		for c in range(3, max_col + 1):
			header = ws.cell(header_row, c).value
			if not isinstance(header, str):
				continue
			stype = _header_to_standard_type(header.strip())
			if stype is not None:
				type_columns.append((c, stype))

		if not type_columns:
			raise CommandError(
				"No statement-type columns found in the header row. "
				"Expected headers containing keywords like 'Statement Text', "
				"'needs attention', 'strong standard', 'exceptional', 'urgent improvement'."
			)

		self.stdout.write(f"Header row: {header_row}")
		self.stdout.write("Detected statement columns:")
		for col_idx, stype in type_columns:
			header_text = ws.cell(header_row, col_idx).value
			self.stdout.write(f"  Col {col_idx} ({header_text!r}) → {stype}")

		# ── Collect all statements ───────────────────────────────────────────
		area_order: dict[str, int] = {}
		# (area_name, statement_number, text, standard_type)
		statements: list[tuple[str, int, str, InDepthStatement.StandardType]] = []

		for r in range(header_row + 1, max_row + 1):
			area_val = ws.cell(r, 1).value
			num_val = ws.cell(r, 2).value

			# Skip fully blank rows
			if area_val is None and num_val is None:
				continue
			if not area_val:
				continue

			area_name = str(area_val).strip()
			if not area_name:
				continue

			try:
				statement_number = int(num_val)
			except (TypeError, ValueError):
				continue

			if area_name not in area_order:
				area_order[area_name] = len(area_order) + 1

			# One InDepthStatement per type column that has real content
			for col_idx, stype in type_columns:
				cell_val = ws.cell(r, col_idx).value
				if _is_skippable(cell_val):
					continue
				text = str(cell_val).strip()
				if text:
					statements.append((area_name, statement_number, text, stype))

		# ── Write to database ────────────────────────────────────────────────
		created_areas = updated_areas = created_statements = updated_statements = 0

		with transaction.atomic():
			areas_by_name: dict[str, InDepthArea] = {}
			for name, order in area_order.items():
				obj, created = InDepthArea.objects.update_or_create(
					name=name,
					defaults={"order": order},
				)
				areas_by_name[name] = obj
				if created:
					created_areas += 1
				else:
					updated_areas += 1

			for area_name, statement_number, statement_text, standard_type in statements:
				area_obj = areas_by_name[area_name]
				obj, created = InDepthStatement.objects.update_or_create(
					area=area_obj,
					standard_type=standard_type,
					statement_number=statement_number,
					defaults={"text": statement_text},
				)
				if created:
					created_statements += 1
				else:
					updated_statements += 1

		self.stdout.write(
			self.style.SUCCESS(
				f"Done. Areas: {created_areas} created, {updated_areas} updated. "
				f"Statements: {created_statements} created, {updated_statements} updated."
			)
		)
