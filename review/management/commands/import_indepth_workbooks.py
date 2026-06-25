"""
Import in-depth review criteria directly from the supporting-tool workbooks
(the per-area `.xlsx` files Oxlip issues, e.g. "Updated Achievement 2026.xlsx").

Builds:  InDepthArea -> InDepthStandard (grade band) -> InDepthJudgementArea

One workbook == one evaluation area, with six sheets:
    Guidance, Expected Standard, Strong Standard,
    Urgent Improvement, Needs Attention, Exceptional

Columns are located by *header text*, not position, so per-tool variations are
handled automatically — notably Leadership & Governance's extra
"How do we know this?" column, which other tools omit.

The import is idempotent: an area's standards are upserted and each standard's
judgement areas are replaced wholesale, so a revised drop of a workbook can be
re-run safely. Only areas whose workbooks are present are touched; everything
else (loaded from criteria.json) is left intact. Next week's remaining tools
drop into the same folder and reuse this exact path.

Run with:
    python manage.py import_indepth_workbooks                 # default folder
    python manage.py import_indepth_workbooks --dir <folder>  # custom folder
    python manage.py import_indepth_workbooks --path <file>   # single workbook
"""
from __future__ import annotations

import re
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from review.models import (
    InDepthArea,
    InDepthJudgementArea,
    InDepthStandard,
)

try:
    import openpyxl
except ImportError:  # pragma: no cover - dependency is declared in requirements
    openpyxl = None

# Sheet title -> InDepthStandard.Key value (Guidance carries no criteria).
SHEET_TO_KEY = {
    "Expected Standard": InDepthStandard.Key.EXPECTED_STANDARD,
    "Strong Standard": InDepthStandard.Key.STRONG_STANDARD,
    "Urgent Improvement": InDepthStandard.Key.URGENT_IMPROVEMENT,
    "Needs Attention": InDepthStandard.Key.NEEDS_ATTENTION,
    "Exceptional": InDepthStandard.Key.EXCEPTIONAL,
    "Met": InDepthStandard.Key.MET,
    "Not Met": InDepthStandard.Key.NOT_MET,
}

# The two "rich" sheets carry full judgement-area blocks; the rest are flat lists.
RICH_KEYS = {
    InDepthStandard.Key.EXPECTED_STANDARD,
    InDepthStandard.Key.STRONG_STANDARD,
}

# Display/import order of the standards within an area.
KEY_ORDER = {
    InDepthStandard.Key.URGENT_IMPROVEMENT: 1,
    InDepthStandard.Key.NEEDS_ATTENTION: 2,
    InDepthStandard.Key.EXPECTED_STANDARD: 3,
    InDepthStandard.Key.STRONG_STANDARD: 4,
    InDepthStandard.Key.EXCEPTIONAL: 5,
    InDepthStandard.Key.NOT_MET: 1,
    InDepthStandard.Key.MET: 2,
}

# Flat-shape standards whose statements are nonetheless RAG-able rungs in the
# ladder (Urgent Improvement on the down-path, Exceptional on the up-path).
# Needs Attention is only an outcome label, so it stays reference-only.
RATEABLE_FLAT_KEYS = {
    InDepthStandard.Key.URGENT_IMPROVEMENT,
    InDepthStandard.Key.EXCEPTIONAL,
}

DEFAULT_DIR = Path(settings.BASE_DIR) / "review" / "data" / "workbooks"


def _clean(value) -> str:
    """Normalise a cell to a trimmed string ('' for blanks)."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _norm_header(value) -> str:
    """Header key for matching: lowercased, trailing colon/space removed."""
    return _clean(value).rstrip(":").strip().lower()


class Command(BaseCommand):
    help = "Import in-depth review criteria from the per-area .xlsx supporting-tool workbooks."

    def add_arguments(self, parser):
        parser.add_argument(
            "--dir",
            default=str(DEFAULT_DIR),
            help="Folder of .xlsx workbooks (defaults to review/data/workbooks).",
        )
        parser.add_argument(
            "--path",
            default=None,
            help="Import a single workbook instead of a whole folder.",
        )

    @transaction.atomic
    def handle(self, *args, **opts):
        if openpyxl is None:
            raise CommandError("openpyxl is required to import workbooks (pip install openpyxl).")

        if opts["path"]:
            paths = [Path(opts["path"])]
        else:
            folder = Path(opts["dir"])
            if not folder.exists():
                raise CommandError(f"Workbook folder not found: {folder}")
            paths = sorted(p for p in folder.glob("*.xlsx") if not p.name.startswith("~$"))

        if not paths:
            raise CommandError("No .xlsx workbooks found to import.")

        n_areas = n_standards = n_ja = 0
        for path in paths:
            if not path.exists():
                raise CommandError(f"Workbook not found: {path}")
            a, s, j = self._import_workbook(path)
            n_areas += a
            n_standards += s
            n_ja += j

        self.stdout.write(
            self.style.SUCCESS(
                f"Imported {n_areas} area(s), {n_standards} standard(s), "
                f"{n_ja} judgement area(s)/statement(s) from {len(paths)} workbook(s)."
            )
        )

    # ── per-workbook ────────────────────────────────────────────────────────
    def _import_workbook(self, path: Path):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            area_name = self._area_name(wb)
            if not area_name:
                raise CommandError(f"Could not determine area name in {path.name}.")

            area, _ = InDepthArea.objects.get_or_create(name=area_name)
            self.stdout.write(f"  {path.name} -> area '{area.name}'")

            n_standards = n_ja = 0
            for sheet_title in wb.sheetnames:
                key = SHEET_TO_KEY.get(sheet_title)
                if key is None:
                    continue  # Guidance / unrecognised sheets

                ws = wb[sheet_title]
                if key in RICH_KEYS:
                    focus, jas = self._parse_rich(ws)
                    usage_notes = []
                else:
                    focus, usage_notes, jas = self._parse_flat(ws, key)

                standard, _ = InDepthStandard.objects.update_or_create(
                    area=area,
                    key=key,
                    defaults={
                        "focus": focus,
                        "usage_notes": usage_notes,
                        "order": KEY_ORDER.get(key, 0),
                    },
                )
                n_standards += 1

                standard.judgement_areas.all().delete()
                rows = [
                    InDepthJudgementArea(standard=standard, order=i + 1, **ja)
                    for i, ja in enumerate(jas)
                ]
                InDepthJudgementArea.objects.bulk_create(rows)
                n_ja += len(rows)

            return 1, n_standards, n_ja
        finally:
            wb.close()

    def _area_name(self, wb) -> str:
        """Area name lives in column A of the header row of any standard sheet."""
        for title in ("Expected Standard", "Strong Standard"):
            if title in wb.sheetnames:
                # Rich sheets: header on row 2, area name in A2.
                ws = wb[title]
                val = _clean(ws["A2"].value)
                if val:
                    return val
        # Fallback: any flat sheet (header on row 1, area name in A1).
        for title in wb.sheetnames:
            if title in SHEET_TO_KEY:
                return _clean(wb[title]["A1"].value)
        return ""

    # ── rich sheet (Expected / Strong) ──────────────────────────────────────
    def _parse_rich(self, ws):
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return "", []

        header = rows[1]  # row 2
        cols = self._rich_columns(header)

        # Focus sits in row 1, under the statement column (blank for some tools).
        focus = ""
        if cols["statement"] is not None and len(rows[0]) > cols["statement"]:
            focus = _clean(rows[0][cols["statement"]])

        def cell(row, name):
            idx = cols[name]
            if idx is None or idx >= len(row):
                return ""
            return _clean(row[idx])

        jas = []
        current = None
        for row in rows[2:]:  # data rows from row 3
            statement = cell(row, "statement")
            if statement:
                current = {
                    "statement": statement,
                    "key_questions": [],
                    "suggested_evidence": [],
                    "how_we_know": [],
                    "sources": [],
                    "example_commentary": cell(row, "commentary"),
                    "example_next_steps": cell(row, "next_steps"),
                    "is_flat": False,
                }
                jas.append(current)
            elif current is None:
                continue  # skip anything before the first statement

            if current is None:
                continue
            for field, header_name in (
                ("key_questions", "key_questions"),
                ("suggested_evidence", "suggested_evidence"),
                ("how_we_know", "how_we_know"),
                ("sources", "sources"),
            ):
                val = cell(row, header_name)
                if val:
                    current[field].append(val)
            # Exemplar text on a continuation row (rare) — keep it.
            if not statement:
                for field, header_name in (
                    ("example_commentary", "commentary"),
                    ("example_next_steps", "next_steps"),
                ):
                    val = cell(row, header_name)
                    if val:
                        current[field] = (current[field] + "\n" + val).strip()

        return focus, jas

    @staticmethod
    def _rich_columns(header):
        """Map logical fields to column indices by matching header text."""
        cols = {
            "statement": None,
            "key_questions": None,
            "suggested_evidence": None,
            "how_we_know": None,
            "sources": None,
            "commentary": None,
            "next_steps": None,
        }
        for idx, raw in enumerate(header):
            h = _norm_header(raw)
            if not h:
                continue
            if "judgement" in h and cols["statement"] is None:
                cols["statement"] = idx
            elif h == "key questions":
                cols["key_questions"] = idx
            elif h == "suggested evidence":
                cols["suggested_evidence"] = idx
            elif h == "how do we know this?" or h == "how do we know this":
                cols["how_we_know"] = idx
            elif h.startswith("where schools might source"):
                cols["sources"] = idx
            elif h == "commentary":
                cols["commentary"] = idx
            elif h == "next steps":
                cols["next_steps"] = idx
        return cols

    # ── flat sheet (Urgent Improvement / Needs Attention / Exceptional) ──────
    def _parse_flat(self, ws, key):
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return "", [], []

        header = rows[0]  # row 1
        statement_col = None
        notes_col = None
        for idx, raw in enumerate(header):
            h = _norm_header(raw)
            if not h:
                continue
            if "judgement" in h and statement_col is None:
                statement_col = idx
            elif statement_col is not None and notes_col is None:
                # The single column after the judgements column carries the
                # sheet's usage guidance ("These are examples…", "Do any of
                # these statements apply?…").
                notes_col = idx
        if statement_col is None:
            statement_col = 1  # fall back to column B

        is_flat = key not in RATEABLE_FLAT_KEYS
        statements = []
        usage_notes = []
        for row in rows[1:]:
            if statement_col < len(row):
                stmt = _clean(row[statement_col])
                if stmt:
                    statements.append(stmt)
            if notes_col is not None and notes_col < len(row):
                note = _clean(row[notes_col])
                if note:
                    usage_notes.append(note)

        jas = [
            {
                "statement": stmt,
                "key_questions": [],
                "suggested_evidence": [],
                "how_we_know": [],
                "sources": [],
                "example_commentary": "",
                "example_next_steps": "",
                "is_flat": is_flat,
            }
            for stmt in statements
        ]
        return "", usage_notes, jas
